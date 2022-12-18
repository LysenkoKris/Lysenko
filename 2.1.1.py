import csv
import math
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side

name_list = ["name", "salary_from", "salary_to", "salary_currency", "area_name", "published_at"]

currency_to_rub = {"AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76,
        "KZT": 0.13, "RUR": 1, "UAH": 1.64, "USD": 60.66, "UZS": 0.0055}

list_print1 = ['Динамика уровня зарплат по годам: ','Динамика количества вакансий по годам: ',
                      'Динамика уровня зарплат по годам для выбранной профессии: ','Динамика количества вакансий по годам для выбранной профессии: ',
                      'Уровень зарплат по городам (в порядке убывания): ','Доля вакансий по городам (в порядке убывания): ']


class Vacancy:
    """Класс для получения данных о вакансии.

    Attributes:
        name (str): Название вакансии
        salary_average (int): Средняя зарплата в рублях
        area_name (str): Название города
        publication_year (int): Год публикации вакансии
    """
    def __init__(self, vacancy):
        """Инициализирует объект Vacancy, вычисляет среднюю зарплату и переводит в рубли

        Args:
            vacancy (dict): Вакансия
        """
        self.name = vacancy[name_list[0]]
        self.salary_average = math.floor((float(vacancy[name_list[1]]) + float(vacancy[name_list[2]])) / 2) \
                              * currency_to_rub[vacancy[name_list[3]]]
        self.area_name = vacancy[name_list[4]]
        self.publication_year = int(vacancy[name_list[5]][:4])


class DataSet:
    """Класс для получения и печати статистик.

    Attributes:
        filename (str): Название файла с данными о вакансиях
        vacancy_name (str): Название выбранной профессии
    """
    def __init__(self, filename, vacancy_name):
        """Инициализирует объект DataSet.

        Args:
            filename (str): Название файла с данными о вакансиях
            vacancy_name (str): Название выбранной профессии
        """
        self.filename, self.vacancy_name = filename, vacancy_name

    def csv_reader(self):
        """Считывает данные из входного файла

        Returns:
            dict: Все вакансии с информацией о них.
        """
        with open(self.filename, mode='r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            header = next(reader)
            header_length = len(header)
            for row in reader:
                if '' not in row and len(row) == header_length:
                    yield dict(zip(header, row))

    @staticmethod
    def average(dict):
        """Высчитывает среднее значение.

        Args:
            dict (dict): Словарь с значениями
        Returns:
            dict: Словарь с обновленными, средними значениями
        """
        new_dict = {}
        for k, v in dict.items():
            new_dict[k] = int(sum(v) / len(v))
        return new_dict

    @staticmethod
    def increment(dict, k, с):
        """Дополняет словарь всех средних зарплат за год или в городе

        Args:
            dict (dict): Словарь со всеми зарплатами за год или в городе по вакансии
            k (int): Год или город вакансии, по которому идет подсчет
            с (list): Средняя зарплата у вакансии
        """
        if k in dict:
            dict[k] += с
        else:
            dict[k] = с

    def get_dynamics(self):
        """Получает все необходимые статистики для дальнейшей работы

        Returns:
            dict, dict, dict, dict, dict, dict: Все необходимые статистики
        """
        salary = {}
        salary_of_name = {}
        city = {}
        count = 0

        for vacancy_dictionary in self.csv_reader():
            vacancy = Vacancy(vacancy_dictionary)
            self.increment(salary, vacancy.publication_year, [vacancy.salary_average])
            if vacancy.name.find(self.vacancy_name) != -1:
                self.increment(salary_of_name, vacancy.publication_year, [vacancy.salary_average])
            self.increment(city, vacancy.area_name, [vacancy.salary_average])
            count += 1

        number = dict([(k, len(v)) for k, v in salary_of_name.items()])
        vacancy_number = dict([(k, len(v)) for k, v in salary.items()])

        if not salary_of_name:
            number = dict([(k, 0) for k, v in vacancy_number.items()])
            salary_of_name = dict([(k, [0]) for k, v in salary.items()])

        dynamics1, dynamics2, dynamics3 = self.average(salary), self.average(salary_of_name), self.average(city)

        dynamics4 = {}
        for y, s in city.items():
            dynamics4[y] = round(len(s) / count, 4)
        dynamics4 = list(filter(lambda x: x[-1] >= 0.01, [(k, v) for k, v in dynamics4.items()]))
        dynamics4.sort(key=lambda x: x[-1], reverse=True)
        dynamics5 = dict(dynamics4.copy()[:10])
        dynamics4 = dict(dynamics4)
        dynamics3 = list(filter(lambda x: x[0] in list(dynamics4.keys()), [(k, v) for k, v in dynamics3.items()]))
        dynamics3.sort(key=lambda x: x[-1], reverse=True)
        dynamics3 = dict(dynamics3[:10])


        return dynamics1, vacancy_number, dynamics2, number, dynamics3, dynamics5

    @staticmethod
    def print_statistic(dynamics1, dynamics2, dynamics3, dynamics4, dynamics5, dynamics6):
        """Выводит все динамики с описанием

        Args:
            dynamics1 (dict): Динамика уровня зарплат по годам
            dynamics2 (dict): Динамика количества вакансий по годам
            dynamics3 (dict): Динамика уровня зарплат по годам для выбранной профессии
            dynamics4 (dict): Динамика количества вакансий по годам для выбранной профессии
            dynamics5 (dict): Уровень зарплат по городам (в порядке убывания)
            dynamics6 (dict): Доля вакансий по городам (в порядке убывания)
        Prints:
            Печатать каждой динамики с описанием
        """
        list_print2 = [dynamics1, dynamics2, dynamics3, dynamics4, dynamics5, dynamics6]
        for i in range(len(list_print1)):
            print(list_print1[i]+'{0}'.format(list_print2[i]))


class InputConnect:
    """Класс для получения объектов DataSet и Report.

    Attributes:
        filename (str): Название файла с данными о вакансиях
        name_vacancy (str): Название выбранной профессии
    """
    def __init__(self):
        """Инициализирует объект InputConnect.
        """
        self.filename, self.name_vacancy = input('Введите название файла: '), input('Введите название профессии: ')

        dataset = DataSet(self.filename, self.name_vacancy)
        dynamics1, dynamics2, dynamics3, dynamics4, dynamics5, dynamics6 = dataset.get_dynamics()
        dataset.print_statistic(dynamics1, dynamics2, dynamics3, dynamics4, dynamics5, dynamics6)

        report = Report(self.name_vacancy, dynamics1, dynamics2, dynamics3, dynamics4, dynamics5, dynamics6)
        report.generate_excel()


class Report:
    """Класс для получения отчета по полученным динамикам.

    Attributes:
        workbook (Workbook()): Рабочая книга для получания xlsx-файла
        name_vacancy (str): Название выбранной профессии
        dynamics1 (dict): Динамика уровня зарплат по годам
        dynamics2 (dict): Динамика количества вакансий по годам
        dynamics3 (dict): Динамика уровня зарплат по годам для выбранной профессии
        dynamics4 (dict): Динамика количества вакансий по годам для выбранной профессии
        dynamics5 (dict): Уровень зарплат по городам (в порядке убывания)
        dynamics6 (dict): Доля вакансий по городам (в порядке убывания)
    """
    def __init__(self, name_vacancy, dynamics1, dynamics2, dynamics3, dynamics4, dynamics5, dynamics6):
        """Инициализирует объект Report и получает объект Workbook.

        Args:
            name_vacancy (str): Название выбранной профессии
            dynamics1 (dict): Динамика уровня зарплат по годам
            dynamics2 (dict): Динамика количества вакансий по годам
            dynamics3 (dict): Динамика уровня зарплат по годам для выбранной профессии
            dynamics4 (dict): Динамика количества вакансий по годам для выбранной профессии
            dynamics5 (dict): Уровень зарплат по городам (в порядке убывания)
            dynamics6 (dict): Доля вакансий по городам (в порядке убывания)
        """
        self.workbook = Workbook()
        self.name_vacancy = name_vacancy
        self.dynamics1, self.dynamics2, self.dynamics3, self.dynamics4, self.dynamics5, self.dynamics6 \
            = dynamics1, dynamics2, dynamics3, dynamics4, dynamics5, dynamics6

    def generate_excel(self):
        """Генерирует две страницы в рабочей книге с статистикой по годам и статистикой по городам, после чего сохраняет рабочую книгу в файл report.xlsx
        """
        work_sheet1 = self.workbook.active
        work_sheet1.title = 'Статистика по годам'
        work_sheet1.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.name_vacancy, 'Количество вакансий', 'Количество вакансий - ' + self.name_vacancy])
        for year in self.dynamics1.keys():
            work_sheet1.append([year, self.dynamics1[year], self.dynamics3[year], self.dynamics2[year], self.dynamics4[year]])

        data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.name_vacancy, ' Количество вакансий', ' Количество вакансий - ' + self.name_vacancy]]
        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]

        for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
            work_sheet1.column_dimensions[get_column_letter(i)].width = column_width + 2

        data = []
        data.append(['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'])
        for (city1, value1), (city2, value2) in zip(self.dynamics5.items(), self.dynamics6.items()):
            data.append([city1, value1, '', city2, value2])
        work_sheet2 = self.workbook.create_sheet('Статистика по городам')
        for row in data:
            work_sheet2.append(row)

        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                cell = str(cell)
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]

        for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
            work_sheet2.column_dimensions[get_column_letter(i)].width = column_width + 2

        bold = Font(bold=True)
        for column in 'ABCDE':
            work_sheet1[column + '1'].font = bold
            work_sheet2[column + '1'].font = bold

        for index, _ in enumerate(self.dynamics5):
            work_sheet2['E' + str(index + 2)].number_format = '0.00%'

        slim = Side(border_style='thin', color='00000000')

        for row in range(len(data)):
            for column in 'ABDE':
                work_sheet2[column + str(row + 1)].border = Border(left=slim, bottom=slim, right=slim, top=slim)

        self.dynamics1[1] = 1
        for row, _ in enumerate(self.dynamics1):
            for column in 'ABCDE':
                work_sheet1[column + str(row + 1)].border = Border(left=slim, bottom=slim, right=slim, top=slim)

        self.workbook.save('report.xlsx')


if __name__ == '__main__':
    InputConnect()