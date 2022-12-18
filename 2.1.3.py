import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side
import matplotlib.pyplot as plt
import numpy as np
from jinja2 import Environment, FileSystemLoader
import pathlib
import pdfkit
import math

currency_to_rub = {"AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76, "KZT": 0.13, "RUR": 1,
                   "UAH": 1.64, "USD": 60.66, "UZS": 0.0055}
name_list = ["name", "salary_from", "salary_to", "salary_currency", "area_name", "published_at"]

list_print1 = ['Динамика уровня зарплат по годам: ','Динамика количества вакансий по годам: ',
                      'Динамика уровня зарплат по годам для выбранной профессии: ','Динамика количества вакансий по годам для выбранной профессии: ',
                      'Уровень зарплат по городам (в порядке убывания): ','Доля вакансий по городам (в порядке убывания): ']


class Vacancy:
    """Класс для получения данных о вакансии.

    Attributes:
        name (str): Название вакансии
        salary_average (int): Средняя зарплата в рублях
        area_name (str): Название города
        publication_year (int): Год публикации вакансии
    """

    def __init__(self, vacancy):
        """Инициализирует объект Vacancy, вычисляет среднюю зарплату и переводит в рубли

        Args:
            vacancy (dict): Вакансия
        """
        self.name = vacancy[name_list[0]]
        self.salary_average = math.floor((float(vacancy[name_list[1]]) + float(vacancy[name_list[2]])) / 2) \
                              * currency_to_rub[vacancy[name_list[3]]]
        self.area_name = vacancy[name_list[4]]
        self.publication_year = int(vacancy[name_list[5]][:4])


class DataSet:
    """Класс для получения и печати статистик.

    Attributes:
        filename (str): Название файла с данными о вакансиях
        name_vacancy (str): Название выбранной профессии
    """
    def __init__(self, filename, name):
        """Инициализирует объект DataSet.

        Args:
            filename (str): Название файла с данными о вакансиях
            name (str): Название выбранной профессии
        """
        self.filename, self.name_vacancy = filename, name

    def csv_reader(self):
        """Считывает данные из входного файла, получает все необходимые статистики для дальнейшей работы и печатает их

        Returns:
            dict, dict, dict, dict, dict, dict: Все необходимые статистики
        """
        with open(self.filename, mode='r', encoding='utf-8-sig') as file:
            count = 0
            salary = {}
            city = {}
            number = {}
            vacancy_number = {}
            salary_of_name = {}
            number_of_name = {}
            header = []
            reader = csv.reader(file)
            for index, row in enumerate(reader):
                if index == 0:
                    csv_header_length = len(row)
                    header = row
                elif '' not in row and len(row) == csv_header_length:
                    vacancies = Vacancy(dict(zip(header, row)))
                    if vacancies.publication_year not in salary:
                        salary[vacancies.publication_year] = [vacancies.salary_average]
                    else:
                        salary[vacancies.publication_year].append(vacancies.salary_average)

                    if vacancies.area_name not in city:
                        city[vacancies.area_name] = [vacancies.salary_average]
                    else:
                        city[vacancies.area_name].append(vacancies.salary_average)

                    if vacancies.area_name in number:
                        number[vacancies.area_name] += 1
                    else:
                        number[vacancies.area_name] = 1

                    if vacancies.publication_year in vacancy_number:
                        vacancy_number[vacancies.publication_year] += 1
                    else:
                        vacancy_number[vacancies.publication_year] = 1

                    if vacancies.name.find(self.name_vacancy) != -1:
                        if vacancies.publication_year not in salary_of_name:
                            salary_of_name[vacancies.publication_year] = [vacancies.salary_average]
                        else:
                            salary_of_name[vacancies.publication_year].append(vacancies.salary_average)

                        if vacancies.publication_year in number_of_name:
                            number_of_name[vacancies.publication_year] += 1
                        else:
                            number_of_name[vacancies.publication_year] = 1
                    count += 1

        if not salary_of_name:
            number_of_name = dict([(k, 0) for k, v in vacancy_number.copy().items()])
            salary_of_name = dict([(k, []) for k, v in salary.copy().items()])

        dynamics1, dynamics2, dynamics3, dynamics4 = {}, {}, {}, {}
        for year, salaries in salary.items():
            dynamics1[year] = int(sum(salaries) / len(salaries))

        for year, salaries in salary_of_name.items():
            dynamics2[year] = 0 if len(salaries) == 0 else int(sum(salaries) / len(salaries))

        for year, salaries in city.items():
            dynamics3[year] = int(sum(salaries) / len(salaries))

        for year, salaries in number.items():
            dynamics4[year] = round(salaries / count, 4)

        dynamics4 = list(filter(lambda a: a[-1] >= 0.01, [(key, value) for key, value in dynamics4.items()]))
        dynamics4.sort(key=lambda a: a[-1], reverse=True)
        dynamics3 = list(
            filter(lambda a: a[0] in list(dict(dynamics4).keys()), [(key, value) for key, value in dynamics3.items()]))
        dynamics3.sort(key=lambda a: a[-1], reverse=True)

        list_print2 = [str(dynamics1), str(vacancy_number), str(dynamics2), str(number_of_name),
                       str(dict(dynamics3[:10])), str(dict(dynamics4.copy()[:10]))]
        for i in range(len(list_print1)):
            print(list_print1[i] + list_print2[i])

        return dynamics1, vacancy_number, dynamics2, number_of_name, dict(dynamics3[:10]), dict(dynamics4.copy()[:10])


class InputConnect:
    """Класс для получения объектов DataSet и Report.

    Attributes:
        filename (str): Название файла с данными о вакансиях
        name (str): Название выбранной профессии
    """
    def __init__(self):
        """Инициализирует объект InputConnect.
        """
        self.filename, self.name = input('Введите название файла: '), input('Введите название профессии: ')

        dataset = DataSet(self.filename, self.name)
        dynamics1, dynamics2, dynamics3, dynamics4, dynamics5, dynamics6 = dataset.csv_reader()
        new_graphic = Report(self.name, dynamics1, dynamics2, dynamics3, dynamics4, dynamics5, dynamics6)
        new_graphic.generate_image()
        work_sheet1 = new_graphic.get_first_sheet()
        work_sheet2, len_new_data = new_graphic.get_second_sheet()
        new_graphic.generate_excel(work_sheet1, work_sheet2, len_new_data)
        new_graphic.generate_pdf()


class Report:
    """Класс для получения отчета по полученным динамикам.

    Attributes:
        workbook (Workbook()): Рабочая книга для получания xlsx-файла
        name_vacancy (str): Название выбранной профессии
        dynamics1 (dict): Динамика уровня зарплат по годам
        dynamics2 (dict): Динамика количества вакансий по годам
        dynamics3 (dict): Динамика уровня зарплат по годам для выбранной профессии
        dynamics4 (dict): Динамика количества вакансий по годам для выбранной профессии
        dynamics5 (dict): Уровень зарплат по городам (в порядке убывания)
        dynamics6 (dict): Доля вакансий по городам (в порядке убывания)
    """
    def __init__(self, name_vacancy, dynamics1, dynamics2, dynamics3, dynamics4, dynamics5, dynamics6):
        """Инициализирует объект Report и получает объект Workbook.

        Args:
            name_vacancy (str): Название выбранной профессии
            dynamics1 (dict): Динамика уровня зарплат по годам
            dynamics2 (dict): Динамика количества вакансий по годам
            dynamics3 (dict): Динамика уровня зарплат по годам для выбранной профессии
            dynamics4 (dict): Динамика количества вакансий по годам для выбранной профессии
            dynamics5 (dict): Уровень зарплат по городам (в порядке убывания)
            dynamics6 (dict): Доля вакансий по городам (в порядке убывания)
        """
        self.workbook = Workbook()
        self.name_vacancy = name_vacancy
        self.dynamics1 = dynamics1
        self.dynamics2 = dynamics2
        self.dynamics3 = dynamics3
        self.dynamics4 = dynamics4
        self.dynamics5 = dynamics5
        self.dynamics6 = dynamics6

    def generate_image(self):
        """Генерирует 4 диаграммы в на одной старнице на основе статистик, после чего сохраняет картинку в файл graph.png
        """
        x = np.arange(len(self.dynamics1.keys()))
        width = 0.35

        fig, axs = plt.subplots(ncols=2, nrows=2)
        axs[0, 0].bar(x - width / 2, self.dynamics1.values(), width, label='средняя з/п')
        axs[0, 0].bar(x + width / 2, self.dynamics3.values(), width, label='з/п {0}'.format(self.name_vacancy))
        plt.rcParams['font.size'] = '8'
        axs[0, 0].set_title('Уровень зарплат по годам')
        axs[0, 0].set_xticks(x, self.dynamics1.keys(), rotation=90)
        axs[0, 0].grid(axis='y')
        axs[0, 0].legend(fontsize=8)

        axs[0, 1].bar(x - width / 2, self.dynamics2.values(), width, label='количество вакансий')
        axs[0, 1].bar(x + width / 2, self.dynamics4.values(), width,
                      label='количество вакансий {0}'.format(self.name_vacancy))
        axs[0, 1].set_title('Количество вакансий по годам')
        axs[0, 1].set_xticks(x, self.dynamics2.keys(), rotation=90)
        axs[0, 1].grid(axis='y')
        axs[0, 1].legend(fontsize=8)
        fig.tight_layout()

        areas = []
        for area in self.dynamics5.keys():
            areas.append(str(area).replace(' ', '\n').replace('-', '-\n'))
        y_pos = np.arange(len(areas))
        performance = self.dynamics5.values()
        error = np.random.rand(len(areas))
        axs[1, 0].barh(y_pos, performance, xerr=error, align='center')
        axs[1, 0].set_yticks(y_pos, labels=areas, size=6)
        axs[1, 0].invert_yaxis()
        axs[1, 0].grid(axis='x')
        axs[1, 0].set_title('Уровень зарплат по городам')

        val = list(self.dynamics6.values()) + [1 - sum(list(self.dynamics6.values()))]
        k = list(self.dynamics6.keys()) + ['Другие']
        axs[1, 1].pie(val, labels=k, startangle=150)
        axs[1, 1].set_title('Доля вакансий по городам')

        plt.tight_layout()
        plt.savefig('graph.png', dpi=300)

    def get_first_sheet(self):
        """Генерирует первую страницу рабочей книги с статистикой по годам.

        Returns:
            Worksheet: Первая страница рабочей книги
        """
        work_sheet1 = self.workbook.active
        work_sheet1.title = 'Статистика по годам'
        work_sheet1.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.name_vacancy, 'Количество вакансий',
                    'Количество вакансий - ' + self.name_vacancy])
        for year in self.dynamics1.keys():
            work_sheet1.append([year, self.dynamics1[year], self.dynamics3[year], self.dynamics2[year], self.dynamics4[year]])

        data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.name_vacancy, ' Количество вакансий',
                 ' Количество вакансий - ' + self.name_vacancy]]
        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]

        for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
            work_sheet1.column_dimensions[get_column_letter(i)].width = column_width + 2
        return work_sheet1

    def get_second_sheet(self):
        """Генерирует вторую страницу рабочей книги с статистикой по городам.

        Returns:
            Worksheet: Вторая страница рабочей книги
            int: Количество строк
        """
        new_data = [['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']]

        for (city1, value1), (city2, value2) in zip(self.dynamics5.items(), self.dynamics6.items()):
            new_data.append([city1, value1, '', city2, value2])
        work_sheet2 = self.workbook.create_sheet('Статистика по городам')
        for r in new_data:
            work_sheet2.append(r)

        column_widths = []
        for r in new_data:
            for i, cell in enumerate(r):
                cell = str(cell)
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]

        for i, column_width in enumerate(column_widths, 1):
            work_sheet2.column_dimensions[get_column_letter(i)].width = column_width + 2

        return work_sheet2, len(new_data)

    def generate_excel(self, work_sheet1, work_sheet2, len_new_data):
        """Сохраняет рабочую книгу в файл report.xlsx
        Args:
            work_sheet1 (Worksheet): Первая страница рабочей книги
            work_sheet2 (Worksheet): Вторая страница рабочей книги
            len_new_data (int): Количество строк
        """
        bold = Font(bold=True)
        for c in 'ABCDE':
            work_sheet1[c + '1'].font = bold
            work_sheet2[c + '1'].font = bold

        for index, _ in enumerate(self.dynamics5):
            work_sheet2['E' + str(index + 2)].number_format = '0.00%'

        slim = Side(border_style='thin', color='00000000')

        for row in range(len_new_data):
            for column in 'ABDE':
                work_sheet2[column + str(row + 1)].border = Border(left=slim, bottom=slim, right=slim, top=slim)

        self.dynamics1[1] = 1
        for row, _ in enumerate(self.dynamics1):
            for column in 'ABCDE':
                work_sheet1[column + str(row + 1)].border = Border(left=slim, bottom=slim, right=slim, top=slim)

        self.workbook.save('report.xlsx')

    def generate_pdf(self):
        """Генирирует и сохраняет файл report.pdf, в котором хранятся report.xlsx и graph.png
        """
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        dynamics = []
        for year in self.dynamics2.keys():
            dynamics.append([year, self.dynamics1[year], self.dynamics2[year], self.dynamics3[year], self.dynamics4[year]])

        for key in self.dynamics6:
            self.dynamics6[key] = round(self.dynamics6[key] * 100, 2)

        pdf_template = template.render({'name': self.name_vacancy,
                                        'path': '{0}/{1}'.format(pathlib.Path(__file__).parent.resolve(), 'graph.png'),
                                        'dynamics': dynamics, 'dynamics5': self.dynamics5, 'dynamics6': self.dynamics6})

        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})


if __name__ == '__main__':
    InputConnect()
